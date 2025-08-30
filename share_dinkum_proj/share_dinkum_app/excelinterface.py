import pandas as pd
from datetime import date
from datetime import datetime
import os
from pathlib import Path
import re

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# Annoying data types
from uuid import UUID
from djmoney.money import Money
from django.db.models.fields.files import FieldFile

import logging
logger = logging.getLogger(__name__)


def get_all_tables_in_excel(filename):
    """A helper function to extract all Named DataTables from an Excel file,
    and store these in a dictionary (key = Table name, value = Dataframe)"""
    wb = load_workbook(filename, data_only=True)
    mapping = {}
    for ws in wb.worksheets:
        for entry, data_boundary in ws.tables.items():
            data = ws[data_boundary]
            content = [[cell.value for cell in ent] for ent in data]
            header = content[0]
            rest = content[1:]
            df = pd.DataFrame(rest, columns=header)


            df = make_tz_naive(df)
            df = df.dropna(how='all')  # Blank rows would cause not null errors.

            mapping[entry] = df
    return mapping


def make_tz_naive(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)
            df[col] = df[col].apply(lambda x: pd.to_datetime(x, errors='coerce').date())
    return df



class ExcelGen:
    def __init__(
        self, title=None, author=None, description=None, template_info=None, url=None
    ):
        """
        Initializes the ExcelGen class with optional parameters for the title, author, and description of the Excel file, as well as template information.

        Args:
            title (str, optional): The title of the Excel file.
            author (str, optional): The author of the Excel file.
            description (str, optional): A description of the Excel file.
            template_info (str, optional): Information about the template used in the Excel file. If not provided, a default string is used.
        """
        self.excel_illegal_characters_re = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        
        self.title = title
        self.author = author
        self.description = description
        self.template_info = (
            template_info
            or r'All data has been added into named tables contained on separate sheets. The table names and accompanying descriptions is described in "table_info". Primary keys, if defined, are denoted with red header cells. This exporting script is available in the following repository: https://MODEC-DnA@dev.azure.com/MODEC-DnA/dna-mops-nemo/_git/datamanagement '
        )
        self.url = url

        self.table_summary = []

        self.wb = Workbook()
        first_worksheet = self.wb.worksheets[0]
        self.wb.remove(first_worksheet)
        #self._add_cover()

        self.sheet_counter = 0

        self.excel_illegal_characters_re = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

        self.id_col_style =  NamedStyle(name="uuid")
        self.id_col_style.alignment = Alignment(shrinkToFit=True)


    def add_table(
        self,
        df,
        table_name,
        description=None,
        pk=None,
        start_row=1,
        start_col=1,
        position_index=None,
        style_map=None,
        width_map=None,
        format_map=None,
        exclude_from_summary=False,
        add_hyperlinks=True,
        value_style_map=None,
    ):
        """
        Adds a table to the Excel workbook using a pandas DataFrame.

        Args:
            df (pd.DataFrame): The pandas DataFrame to be added as a table.
            table_name (str): The name of the table. Max 30 chars.
            description (str, optional): A description of the table.
            pk (str or list, optional): The primary key(s) of the table.
            start_row (int, optional): The starting row for the table. Defaults to 1.
            start_col (int, optional): The starting column for the table. Defaults to 1.
            position_index (int, optional): The position index of the worksheet in the workbook.
            style_map (dict, optional): A dictionary mapping column names to cell styles.
            width_map (dict, optional): A dictionary mapping column names to column widths.
            format_map (dict, optional): A dictionary mapping column names to number formats.
            exclude_from_summary (bool, optional): If True, the table will not be included in the table summary. Defaults to False.
        """


        if style_map is None:
            style_map = {}

        if width_map is None:
            width_map = {}
        if format_map is None:
            format_map = {}
        if pk is None:
            pk = []
        elif type(pk) is str:
            pk = [pk]

        if value_style_map is None:
            value_style_map = {}

        self.sheet_counter += 1


        if not exclude_from_summary:
            self.table_summary.append((self.sheet_counter, table_name, description, len(df)))

        
        df = df.copy()
        

        df = df.reset_index(drop=True)

        cols = df.columns

        # Strip out all nan, NaT etc and replace with None:
        df.astype(object).where(df.notnull(), None)

        if table_name == 'table_info':
            sheet_name = 'Index'
        else:
            sheet_name = f'{self.sheet_counter:02}'


        ws = self.wb.create_sheet(sheet_name, index=position_index)

        for col_index, col in enumerate(cols):
            cell = ws.cell(column=(col_index + start_col), row=start_row)
            cell.value = col

            if col in pk:
                cell.style = 'Accent2'

            cell.alignment = Alignment(vertical='top')

            column_width = width_map.get(col, None)

            if column_width:
                ws.column_dimensions[
                    get_column_letter(col_index + start_col)
                ].width = column_width

        ws.row_dimensions[start_row].height = 32

        for row in df.itertuples():
            for col_index, col in enumerate(cols):
                val_to_print = row[col_index + 1]
                cell = ws.cell(
                    column=(col_index + start_col),
                    row=(row[0] + start_row + 1),
                )

                if isinstance(val_to_print, UUID):
                    val_to_print = str(val_to_print)
                    cell.style = self.id_col_style

                elif isinstance(val_to_print, Money):
                    val_to_print = val_to_print.amount
                
                elif isinstance(val_to_print, FieldFile):
                    val_to_print = str(val_to_print)
                
                elif type(val_to_print) is str:
                    # Illegal unicode characters. 
                    val_to_print = re.sub(self.excel_illegal_characters_re, '', val_to_print)

                    if val_to_print.startswith('=') and not val_to_print.startswith('=HYPERLINK'):
                        # Append apostrophe before leading equals signs to prevent being interpreted as forumla
                        val_to_print = "'" + val_to_print

                    if val_to_print.startswith('http'):
                        cell.hyperlink = val_to_print
                        cell.style = "Hyperlink"
                

                    if val_to_print.startswith('=HYPERLINK'):
                        cell.style = "Hyperlink"

                cell.value = val_to_print

                # removed  cell.alignment = Alignment(vertical='top')


                style_to_apply = style_map.get(col, {})
                if style_to_apply:
                    cell.style = style_to_apply

                if value_style_map:
                    style_to_apply = value_style_map.get(val_to_print, None)
                    if style_to_apply:
                        cell.style = style_to_apply

                number_format = format_map.get(col, None)
                if number_format:
                    cell.number_format = number_format

        table_range = CellRange(
            min_col=start_col,
            min_row=start_row,
            max_col=start_col + len(cols) - 1,
            max_row=start_row + max(len(df), 1), # need at least one row in the table
        )

        tab = Table(displayName=table_name, ref=table_range.coord)

        table_style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        tab.tableStyleInfo = table_style
        ws.add_table(tab)
        ws.freeze_panes = f"A{start_row + 1}"

    def save(self, output_path):
        """
        Saves the Excel workbook to the specified output path.

        Args:
            output_path (str): The path where the Excel workbook will be saved.
        """

        self._add_table_summary()

        for ws in self.wb.worksheets:
            self._autofit_columns(ws)

        self.wb.save(output_path)

    # def _add_cover(self):
    #     metadata_list = [
    #         ('title', self.title),
    #         ('author', self.author),
    #         ('description', self.description),
    #         ('generated_at', date.today().isoformat()),
    #     ]

    #     if self.url:
    #         metadata_list.append(('link', self.url))

    #     if self.template_info:
    #         metadata_list.append(('template_info', self.template_info))

    #     meta_df = pd.DataFrame.from_records(metadata_list, columns=['key', 'value'])

    #     self.add_table(
    #         df=meta_df,
    #         table_name='cover',
    #         pk='key',
    #         width_map={'key': 20, 'value': 100},
    #         exclude_from_summary=True,
    #         value_style_map={self.title: 'Headline 1'},
    #     )

    def _add_table_summary(self):
        table_summary_df = pd.DataFrame.from_records(
            self.table_summary, columns=['sheet_name', 'table_name', 'description', 'num_records']
        )
        table_summary_df['sheet_name'] = table_summary_df['sheet_name'].apply(lambda x : f'{x:02}')
        table_summary_df['link'] = table_summary_df['sheet_name'].apply(lambda x : f'=HYPERLINK("#\'{x}\'!A1", "{x}")')

        if "_table_info" in self.wb:
            ws_to_remove = self.wb["_table_info"]
            self.wb.remove(ws_to_remove)

        self.add_table(
            df=table_summary_df,
            table_name='table_info',
            pk= None, #'table_name',
            width_map={'sheet_name' : 10, 'table_name': 20, 'description': 100, 'num_records': 16},
            position_index=0,
            exclude_from_summary=True,
        )


    def _autofit_columns(self, ws, max_allowable=80):
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column letter

            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    if cell_length > max_allowable:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                except:
                    pass

            ws.column_dimensions[column].width = (
                min(max_length, max_allowable) + 2 * 1.1
            )

